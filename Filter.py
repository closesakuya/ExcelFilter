import openpyxl
import re
import copy
import os
import re
import threading
import time
from typing import List, Tuple
import abc
from datetime import datetime, date, timedelta


class FilterType:
    raw_list = 0
    py_exp = 1
    reg_exp = 2

    @classmethod
    def check(cls, filter_type: int, raw: str, pt: str) -> bool:
        return cls.__check_map[filter_type](raw, pt)

    @staticmethod
    def check_raw_list(raw: str, pt: str) -> bool:
        pt = pt.strip()
        if re.search(r'\t', pt):
            pt = pt.split('\t')
        else:
            pt = pt.split('\n')
        for m_item in pt:
            if raw and str(m_item).strip() == str(raw).strip():
                return True
        return False

    @staticmethod
    def check_reg_exp(raw: str, pt: re.Pattern) -> bool:
        if not raw:
            return False
        if re.search(pt, str(raw).strip()):
            return True
        return False

    @staticmethod
    def check_py_exp(X: str, pt: str) -> bool:
        if not X:
            return False
        # exp = pt.replace('X', str(raw).strip())
        exp = pt
        try:
            ret = eval(exp)
            return ret
        except Exception as e:
            # print(e)
            raise Exception("表达式[{0}]错误,ERROR: {1}".format(exp, e))
        return False

    __check_map = {
        raw_list: lambda x, y: FilterType.check_raw_list(x, y),
        py_exp: lambda x, y: FilterType.check_py_exp(x, y),
        reg_exp: lambda x, y: FilterType.check_reg_exp(x, y)
    }


class _MetaInputTab(metaclass=abc.ABCMeta):
    def __init__(self, file_path: str, title_pos: Tuple, data_pos: Tuple, **kwargs):
        self.__filter_pool = []
        self.title_pos = title_pos
        self.data_pos = data_pos
        self.__paras = kwargs
        self.__title = self.read_title()
        self.__title_map = {}
        self.__read_cnt = 0
        self.__total_cnt = self.get_total_data_len()
        for i, m_item in enumerate(self.__title):
            self.__title_map[m_item] = i

    @abc.abstractmethod
    def read_nxt_raw_data_line(self) -> List:
        raise NotImplementedError

    @abc.abstractmethod
    def get_total_data_len(self) -> int:
        raise NotImplementedError

    @abc.abstractmethod
    def read_title(self) -> List:
        raise NotImplementedError

    @abc.abstractmethod
    def close(self):
        raise NotImplementedError

    def get_progress(self) -> float:
        return self.__read_cnt / self.__total_cnt

    def _read_nxt_raw_data_line(self) -> List:
        fin = self.read_nxt_raw_data_line()
        if fin:
            self.__read_cnt += 1
        return fin

    def add_filter(self, filter_type: int, pattern: str, filter_title: str):
        if filter_type == FilterType.reg_exp:
            pattern = re.compile(pattern)
        self.__filter_pool.append((filter_type, pattern, filter_title))

    def read_nxt_data_line(self) -> List:
        line = self._read_nxt_raw_data_line()
        while line:
            match = True
            for m_item in self.__filter_pool:
                filter_type, pattern, filter_title = m_item
                if self.__title_map.get(filter_title, None) is None:
                    return []
                if not FilterType.check(filter_type, line[self.__title_map[filter_title]], pattern):
                    match = False
                    break
            if match:
                return line
            else:
                line = self._read_nxt_raw_data_line()

        return []


class _MetaOutputTab(metaclass=abc.ABCMeta):
    def __init__(self, file_path: str, title: List[str], filter_title: List[str] = None, **kwargs):
        self.file_path = file_path
        self.__filter_title = filter_title
        if self.__filter_title:
            self.__filter_index_lst = []
            for item in filter_title:
                if item in title:
                    self.__filter_index_lst.append(title.index(item))
        else:
            self.__filter_index_lst = []

    def write(self, raw: List):
        if not self.__filter_index_lst:
            return self.write_raw(raw)
        return self.write_raw([raw[i] for i in self.__filter_index_lst])

    @abc.abstractmethod
    def write_raw(self, raw: List):
        raise NotImplementedError

    @abc.abstractmethod
    def get_cur_len(self) -> int:
        raise NotImplementedError

    @abc.abstractmethod
    def save(self):
        raise NotImplementedError


class ExcelInputTab(_MetaInputTab):
    def __init__(self, file_path: str, title_pos: Tuple, data_pos: Tuple, **kwargs):

        self.__file_path = file_path
        self.__output = None
        self.__tab = None
        try:
            rb = openpyxl.open(file_path, data_only=True)
            self.__output = rb
            self.__tab = self.__output.active
        except Exception as e:
            print("open failed  ", self.__file_path, e)
            raise e

        self.__my_title = []
        self.__max_column = 0
        if self.__tab:
            cur_row = title_pos[0]
            cur_col = title_pos[1]
            self.__max_column = self.__tab.max_column
            if title_pos[0] < 0:
                self.__my_title = [str(i + 1) for i in list(range(self.__max_column))]
            else:
                m_item = self.__tab.cell(cur_row + 1, cur_col + 1).value
                while cur_col < self.__max_column:
                    self.__my_title.append(str(m_item))
                    cur_col += 1
                    m_item = self.__tab.cell(cur_row + 1, cur_col + 1).value

        self.__cur = data_pos[0] + 1
        super().__init__(file_path, title_pos, data_pos, **kwargs)

    def read_nxt_raw_data_line(self) -> List:
        fin = []

        cur_col = self.data_pos[1] + 1
        if self.__tab:
            if self.__cur > self.__tab.max_row:
                return fin
            m_item = self.__tab.cell(self.__cur, cur_col).value
            while cur_col <= self.__max_column:
                fin.append(m_item)
                cur_col += 1
                m_item = self.__tab.cell(self.__cur, cur_col).value
        self.__cur += 1
        return fin

    def get_total_data_len(self) -> int:
        if self.__tab:
            return self.__tab.max_row - self.data_pos[0]
        return 0

    def read_title(self) -> List:
        return self.__my_title

    def close(self):
        if self.__output:
            self.__output.close()


class CsvInputTab(_MetaInputTab):
    def __init__(self, file_path: str, title_pos: Tuple, data_pos: Tuple, **kwargs):

        self.__file_path = file_path
        self.__sep = kwargs.get("sep", ',')
        self.__f = None
        try:
            self.__f = open(file_path, 'r+')
        except FileNotFoundError:
            print("not found ", self.__file_path)
        cnt = 0
        title = []
        if title_pos[0] < 0:
            max_col_len = 0
            try:
                with open(self.__file_path, 'r') as f:
                    max_col_len = f.readline().split(self.__sep).__len__()
            except FileNotFoundError:
                print("not found ", self.__file_path)
            self.__my_title = [str(i + 1) for i in list(range(max_col_len))]
        else:
            if self.__f:
                while cnt < title_pos[0]:
                    self.__f.readline()
                    cnt += 1
                title = self.__f.readline().strip().split(self.__sep)[title_pos[1]:]
                cnt += 1
            self.__my_title = title

        # 读到数据起始
        while cnt < data_pos[0]:
            self.__f.readline()
            cnt += 1

        super().__init__(file_path, title_pos, data_pos, **kwargs)

    def read_nxt_raw_data_line(self) -> List:
        fin = []
        if self.__f:
            fin = self.__f.readline().strip().split(self.__sep)[self.data_pos[1]:]
        return fin

    def get_total_data_len(self) -> int:
        tt_len = 0
        try:
            with open(self.__file_path, 'r') as f:
                tt_len = f.readlines().__len__() - self.data_pos[0]
        except FileNotFoundError:
            print("not found ", self.__file_path)
        return tt_len

    def read_title(self) -> List:
        return self.__my_title

    def close(self):
        if self.__f:
            self.__f.close()


class ExcelOutputTab(_MetaOutputTab):
    def __init__(self, file_path: str, title: List[str], filter_title: List[str] = None, **kwargs):
        super().__init__(file_path, title, filter_title, **kwargs)
        self.__output = openpyxl.Workbook()
        self.__tab = self.__output.active
        self.__cur = 1
        print(filter_title)
        for i, m_item in enumerate(filter_title if filter_title else title):
            self.__tab.cell(self.__cur, i + 1, m_item)
        self.__cur += 1

    def write_raw(self, raw: List):
        for i, m_item in enumerate(raw):
            self.__tab.cell(self.__cur, i + 1, m_item)
        self.__cur += 1

    def get_cur_len(self) -> int:
        return self.__cur

    def save(self):
        print(self.file_path)
        self.__output.save(self.file_path)


class CsvOutputTab(_MetaOutputTab):
    def __init__(self, file_path: str, title: List[str], filter_title: List[str] = None, **kwargs):
        super().__init__(file_path, title, filter_title, **kwargs)
        self.__f = open(file_path, 'w+', encoding='utf-8')
        self.__sep = kwargs.get("sep", ',')
        self.__f.write(self.__sep.join(filter_title if filter_title else title))
        self.__f.write('\n')
        self.__cur = 0

    def write_raw(self, raw: List):
        self.__f.write(self.__sep.join([str(m_item) if m_item is not None else '' for m_item in raw]))
        self.__f.write('\n')
        self.__cur += 1

    def get_cur_len(self) -> int:
        return self.__cur

    def save(self):
        if self.__f:
            self.__f.close()


class Task:
    def __init__(self, tab_in: _MetaInputTab, tab_out: _MetaOutputTab):
        self._i = tab_in
        self._o = tab_out
        self._is_done = False
        self.fault_msg = ""

    def start(self):
        def _t():
            try:
                item = self._i.read_nxt_data_line()
                st = time.time()
                while item and time.time() - st <= 120:
                    self._o.write(item)
                    item = self._i.read_nxt_data_line()

                self._i.close()
                self._o.save()
                self._is_done = True
            except Exception as e:
                print(e)
                self._is_done = True
                self.fault_msg = str(e)

        t = threading.Thread(target=_t)
        t.setDaemon(True)
        t.start()

    def is_done(self):
        return self._is_done

    def get_progress(self):
        return self._i.get_progress()

    def get_write_len(self) -> int:
        return self._o.get_cur_len()


def get_input_tab_by_filename(file_name: str):
    name = os.path.splitext(file_name)[-1].lower()
    r_map = {
        '.csv': CsvInputTab,
        '.xlsx': ExcelInputTab,
        '.xls': ExcelInputTab
    }
    return r_map.get(name, CsvInputTab)


def get_output_tab_by_filename(file_name: str):
    name = os.path.splitext(file_name)[-1].lower()
    r_map = {
        '.csv': CsvOutputTab,
        '.xlsx': ExcelOutputTab,
        '.xls': ExcelOutputTab
    }
    return r_map.get(name, CsvOutputTab)


# if __name__ == "__main__":
    # ret = FilterType.check(FilterType.reg_exp, 'A12', r"[0-9]{2,3}")
    # print(ret)

    # fac = get_input_tab_by_filename("test.csv")
    # a = fac("test.csv", data_pos=(2,1), title_pos=(0,1))
    # print(a.read_nxt_data_line(), a.get_progress())
    # print(a.read_nxt_data_line(), a.get_progress())
    # print(a.read_nxt_data_line(), a.get_progress())
    # a.close()
    # fac = get_input_tab_by_filename("test.xlsx")
    # a = fac("test.xlsx", data_pos=(1,0), title_pos=(0,0))
    # ret = []
    # ret.append(a.read_nxt_data_line())
    # ret.append(a.read_nxt_data_line())
    # ret.append(a.read_nxt_data_line())
    # print(ret)
    # a.close()
    #
    # a_w = ExcelOutputTab("rest.xlsx", title=[str(i) for i in range(20)])
    # for m_item in ret:
    #     a_w.write(m_item)
    # a_w.save()


